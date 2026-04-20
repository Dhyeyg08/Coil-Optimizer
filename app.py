from flask import Flask, request, render_template_string, jsonify, send_file

import io

import math

from openpyxl import Workbook

from openpyxl.styles import Font, Alignment

app = Flask(__name__)

SCALE = 10  # 0.1 mm precision

# =====================================================

# 🔥 EXACT DEMAND-DRIVEN PLANNER (WITH DEBUG)

# =====================================================

def exact_plan(order, master_width, coil_weight, tolerance, min_util):

    weight_per_mm = (coil_weight * 1000) / master_width

    sizes = [s for s, _ in order]
    demands = {s: d for s, d in order}

    wps_map = {s: (s * weight_per_mm) / 1000 for s in sizes}

    best_plan = None
    best_score = -1e18

    # max slits possible per size (safe upper bound)
    max_slits = {
        s: int((demands[s] + tolerance) / wps_map[s]) + 2
        for s in sizes
    }

    def backtrack(i, current_plan, used_width, produced):

        nonlocal best_plan, best_score

        # ❌ width overflow
        if used_width > master_width:
            return

        # if all sizes processed
        if i == len(sizes):

            util = used_width / master_width
            if util < min_util:
                return

            # tolerance check
            for s in sizes:
                if produced[s] > demands[s] + tolerance:
                    return

            # 🎯 scoring
            width_score = used_width
            demand_score = 0

            for s in sizes:
                demand_score -= abs(produced[s] - demands[s])

            score = width_score * 1000 + demand_score * 100

            if score > best_score:
                best_score = score
                best_plan = current_plan.copy()

            return

        size = sizes[i]
        wps = wps_map[size]

        for slits in range(max_slits[size] + 1):

            new_width = used_width + slits * size

            new_produced = produced.copy()
            new_produced[size] += slits * wps

            current_plan[size] = slits

            backtrack(i + 1, current_plan, new_width, new_produced)

    # start
    backtrack(
        0,
        {s: 0 for s in sizes},
        0,
        {s: 0 for s in sizes}
    )

    if not best_plan:
        return []

    # build result
    result = []
    used_width = 0
    total_weight = 0

    for s in sizes:
        slits = best_plan[s]
        if slits == 0:
            continue

        wps = wps_map[s]
        width = slits * s
        total = slits * wps

        used_width += width
        total_weight += total

        result.append({
            "size": s,
            "slits": slits,
            "width": round(width, 2),
            "weight_per_mm": round(weight_per_mm, 2),
            "weight_per_slit": round(wps, 3),
            "total_weight": round(total, 3)
        })

    remaining = master_width - used_width
    util = used_width / master_width

    return [{
        "coil": result,
        "used_width": round(used_width, 2),
        "remaining_width": round(remaining, 2),
        "utilization": round(util, 4),
        "total_weight": round(total_weight, 3)
    }]

# =====================================================
# 🎨 FRONTEND
# =====================================================
HTML = """
<!DOCTYPE html>
<html>
<head>
<title>Coil Optimizer</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>

<style>
body {
  font-family:'Segoe UI';
  background:#f8fafc;
  color:#1e293b;
  padding:20px;
}

.container {
  max-width:900px;
  margin:auto;
}

.card {
  background:white;
  padding:20px;
  border-radius:12px;
  margin-bottom:20px;
  box-shadow: 0 4px 12px rgba(0,0,0,0.05);
}

input {
  padding:10px;
  margin:6px;
  border-radius:6px;
  border:1px solid #e2e8f0;
  width:140px;
}

button {
  padding:10px 15px;
  border-radius:6px;
  border:none;
  cursor:pointer;
}

.btn-primary {background:#22c55e;color:white;}
.btn-secondary {background:#3b82f6;color:white;}
.btn-danger {background:#ef4444;color:white;}

table {
  width:100%;
  border-collapse:collapse;
  margin-top:10px;
}

th {
  font-weight:700;
  background:#f1f5f9;
}

th,td {
  padding:10px;
  border-bottom:1px solid #e2e8f0;
  text-align:center;
}

.coil-card {
  background:white;
  padding:15px;
  margin-top:15px;
  border-radius:10px;
  box-shadow: 0 2px 10px rgba(0,0,0,0.04);
}

.total-text {
  font-weight:700;
  color:#16a34a;
}

.input-wrapper {
  position: relative;
  display: inline-block;
}

.input-wrapper input {
  padding: 10px 40px 10px 10px;
  border-radius: 8px;
  border: 1px solid #e2e8f0;
}

.input-wrapper span {
  position: absolute;
  right: 10px;
  top: 50%;
  transform: translateY(-50%);
  color: #64748b;
  font-size: 13px;
}
</style>
</head>
<body>
<div class="container">
<h1>Coil Optimization System</h1>

<div class="card">
<div class="input-wrapper">
  <input id="master_width" placeholder="Master Width">
  <span>mm</span>
</div>

<div class="input-wrapper">
  <input id="coil_weight" placeholder="Coil Weight (MT)">
  <span>MT</span>
</div>

<div class="input-wrapper">
  <input id="tolerance" placeholder="Tolerance (Kg)">
  <span>Kg</span>
</div>

<div class="input-wrapper">
  <input id="min_utilization" placeholder="Min Utilization %">
  <span>%</span>
</div>
</div>

<div class="card">
<table id="orderTable">
<tr><th>Size (mm)</th><th>Weight (MT)</th><th>Action</th></tr>
</table>
<button class="btn-secondary" onclick="addRow()">+ Add Row</button>
</div>

<button class="btn-primary" onclick="generate()">Generate Plan</button>

<div id="output"></div>

<button class="btn-secondary" onclick="downloadExcel()">Download Excel</button>
</div>

<script>
function addRow(size="", weight="") {
  const table = document.getElementById("orderTable");
  const row = table.insertRow();

  row.innerHTML = `
    <td><input value="${size}"> mm</td>
    <td><input value="${weight}"> MT</td>
    <td><button onclick="this.parentElement.parentElement.remove()">X</button></td>
  `;
}
addRow(); addRow();

async function generate(){
  const rows=document.querySelectorAll("#orderTable tr");
  let order=[];

  rows.forEach((row,i)=>{
    if(i===0)return;
    const inputs=row.querySelectorAll("input");
    const s=parseFloat(inputs[0].value);
    const w=parseFloat(inputs[1].value);
    if(!isNaN(s)&&!isNaN(w)) order.push([s,w]);
  });

  const payload={
    master_width:document.getElementById("master_width").value,
    coil_weight:document.getElementById("coil_weight").value,
    tolerance:document.getElementById("tolerance").value,
    min_utilization:document.getElementById("min_utilization").value,
    order:order
  };

  const res=await fetch("/plan",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify(payload)});
  const data=await res.json();
  window.latestPlans=data;

  render(data);
}

function render(data){
  const out=document.getElementById("output");
  out.innerHTML="<h2>Results</h2>";

  data.forEach((c,i)=>{
    let total = c.coil.reduce((a,b)=>a+b.total_weight,0);

    let html=`<div class="coil-card">
    <h3>Coil ${i+1}</h3>
    <p>Used Width: <b>${c.used_width} mm</b></p>
    <p>Remaining: <b>${c.remaining_width} mm</b></p>
    <p class="total-text">Total Coil Weight: ${total.toFixed(2)} MT</p>

    <table>
    <tr>
    <th>Size</th><th>Slits</th><th>Width</th><th>Wt/mm</th><th>Wt/Slit</th><th>Total</th>
    </tr>`;

    c.coil.forEach(it=>{
      html+=`<tr>
<td>${it.size} mm</td>
<td>${it.slits}</td>
<td>${it.width} mm</td>
<td>${it.weight_per_mm} kg</td>
<td>${it.weight_per_slit} MT</td>
<td>${it.total_weight} MT</td>
</tr>`;
    });

    html+="</table></div>";
    out.innerHTML+=html;
  });
}

async function downloadExcel(){
  const res=await fetch("/export",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify(window.latestPlans)});
  const blob=await res.blob();
  const a=document.createElement("a");
  a.href=URL.createObjectURL(blob);
  a.download="coil_plan.xlsx";
  a.click();
}
</script>
</body>
</html>
"""






# =====================================================

# ROUTES

# =====================================================

@app.route("/")

def home():

    return render_template_string(HTML)

@app.route("/plan", methods=["POST"])

def plan_api():

    d = request.json

    print("\n========== API CALLED ==========")

    print("Payload:", d)

    result = exact_plan(

        d["order"],

        float(d["master_width"]),

        float(d["coil_weight"]),

        float(d["tolerance"]) / 1000,

        float(d["min_utilization"]) / 100

    )

    print("========== RESPONSE ==========")

    print(result)

    print("================================\n")

    if result:

        return jsonify(result)

    return jsonify({"error": "Exact plan not feasible"})

@app.route("/export", methods=["POST"])

def export():

    data = request.json

    wb = Workbook()

    ws = wb.active

    headers = ["Coil","Size(mm)","Slits","Width(mm)","Weight(MT)"]

    ws.append(headers)

    for c in ws[1]:

        c.font = Font(bold=True)

    tw, twt = 0,0

    for i,c in enumerate(data):

        for it in c["coil"]:

            tw += it["width"]

            twt += it["total_weight"]

            ws.append([i+1,it["size"],it["slits"],it["width"],it["total_weight"]])

    ws.append(["TOTAL","","",round(tw,2),round(twt,2)])

    for c in ws[ws.max_row]:

        c.font = Font(bold=True)

    for row in ws.iter_rows():

        for c in row:

            c.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()

    wb.save(buf)

    buf.seek(0)

    return send_file(buf, download_name="coil_plan.xlsx", as_attachment=True)

# =====================================================

# RUN

# =====================================================

if __name__ == "__main__":

    app.run(debug=True)
